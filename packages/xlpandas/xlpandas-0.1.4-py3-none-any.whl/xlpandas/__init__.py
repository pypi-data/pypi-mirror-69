'''
Wrapper around Pandas and openpyxl reading and writing data in Excel xlsx/xlst workbooks.
Preserves workbook formatting and structure by only accessing data in rows using openpyxl.

Example
```
import xlpandas as xpd

# Read excel file
df = xpd.read_file('template.xlsx', skiprows=2)
print(df.columns)

# Access openpyxl worksheet
sheet = df.to_sheet()
sheet.cell(1,1).value = 'title'

# From openpyxl worksheet
df = xpd.xlDataFrame(sheet)

# Write file
df.to_file('out.xlsx')
```

'''

import pandas as pd
import openpyxl
from io import BytesIO

class xlDataFrame(pd.DataFrame):
    '''
    Create a DataFrame from a openpyxl worksheet or arraylike.

    xlDataFrame(worksheet=<openpyxl.worksheet>, skiprows=0, **kwargs)
    xlDataFrame(<arraylike>, **kwargs)

    int skiprow - Number of rows to skip before reading table headers

    kwargs passed to pandas DataFrame constructor.
    '''
    _metadata = ["_worksheet", "_skiprows"]
    def __init__(self, *args, **kwargs):
        '''Create dataframe from worksheet'''
        if args and isinstance(args[0], openpyxl.worksheet.worksheet.Worksheet):
            worksheet = args[0]
            args = args[1:]
        else:
            worksheet = kwargs.pop('worksheet', None)
        self._skiprows = kwargs.pop('skiprows', 0) + 1 # always skip header row

        if worksheet is None:
            super().__init__(*args, **kwargs)
            try:
                self.__getattr__('_worksheet')
            except:
                workbook = openpyxl.Workbook()
                self._worksheet = workbook.active
        else:
            self._worksheet = worksheet
            values = [r for r in worksheet.values]
            if kwargs.get('columns') is None and len(values):
                kwargs['columns'] = values[self._skiprows-1]
            super().__init__(values[self._skiprows:], *args, **kwargs)

    @property
    def _constructor(self):
        return xlDataFrame

    # From geopandas
    def __finalize__(self, other, method=None, **kwargs):
        """propagate metadata from other to self """
        # merge operation: using metadata of the left object
        if method == "merge":
            for name in self._metadata:
                object.__setattr__(self, name, getattr(other.left, name, None))
        # concat operation: using metadata of the first object
        elif method == "concat":
            for name in self._metadata:
                object.__setattr__(self, name, getattr(other.objs[0], name, None))
        else:
            for name in self._metadata:
                object.__setattr__(self, name, getattr(other, name, None))
        return self

    def _commit(self, worksheet=None):
        '''Sync DataFrame to worksheet. Excess rows and columns deleted'''
        worksheet = worksheet or self._worksheet
        skiprows = self._skiprows
        h, w = self.shape

        for c in range(w):
            worksheet.cell(skiprows, c+1, self.columns[c])
            for r in range(h):
                value = self.iloc[r, c]
                if pd.isna(value):
                    value = None
                worksheet.cell(skiprows+r+1, c+1).value = value

        excessrows = worksheet.max_row - skiprows - h
        excesscols = worksheet.max_column - w
        if excessrows:
            worksheet.delete_rows(worksheet.max_row - excessrows + 1, excessrows)
        if excesscols:
            worksheet.delete_rows(worksheet.max_columns - excesscols + 1, excesscols)
        return self

    def _copy_sheet(self):
        '''Deep copy workbook'''
        # there is no deepcopy method in openpyxl currently...
        bytestring = BytesIO()
        self._worksheet.parent.save(bytestring)
        return openpyxl.open(bytestring)[self._worksheet.title]

    def copy(self):
        '''Copy DataFrame and worksheet object'''
        df = xlDataFrame(super().copy(), skiprows=self._skiprows)
        df._worksheet = self.as_sheet()
        return df

    def as_sheet(self):
        '''Returns a copy of the worksheet object'''
        newsheet = self._copy_sheet()
        self._commit(newsheet)
        return newsheet

    def to_file(self, filename):
        '''Save workbook to file name'''
        return self.as_sheet().parent.save(filename)

def read_file(filename, sheet=0, **kwargs):
    '''Read worksheet from xlsx/xlst workbook

    string filename - path to xlsx workbook
    string|int sheet - index or name of sheet to read from workbook

    kwargs passed to pandas DataFrame constructor

    returns xlDataFrame
    '''
    workbook = openpyxl.open(filename)
    if type(sheet) is int:
        _sheet = workbook.worksheets[sheet]
    else:
        _sheet = workbook[sheet]
    return xlDataFrame(_sheet, **kwargs)
