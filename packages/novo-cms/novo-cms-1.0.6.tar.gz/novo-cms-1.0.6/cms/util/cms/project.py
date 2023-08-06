# -*- encoding: utf8 -*-
"""
    CMS SubProject Crawler
"""
import os
import re
import sys
import time
import json
import math
import logging
import textwrap

from collections import defaultdict

import openpyxl

import requests

from .title import title, stage_title
from . import parse_config, parse_report_data, write_title

from cms.util import get_logger, User

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.realpath(__file__))))

DEFAULT_CONFIGS = [
    './cms.config.xlsx',
    os.path.join(os.path.expanduser('~'), 'cms.config.xlsx'),
    os.path.join(ROOT_DIR, 'conf', 'cms.config.xlsx'),
]

for each in DEFAULT_CONFIGS:
    if os.path.isfile(each):
        DEFAULT_CONFIG = each
        break


reload(sys)
sys.setdefaultencoding('utf-8')


class CMS(object):
    base_url = 'http://cms.novogene.com'

    def __init__(self, username, password, logger=None):
        self.username = username
        self.password = password
        self.logger = logger or get_logger()
        self.session = requests.session()

    def login(self):
        url = self.base_url + '/core/login/login!login.action'
        formdata = {
            'loginInfo.usercode': self.username,
            'loginInfo.userpass': self.password,
        }
        resp = self.session.post(url, data=formdata)
        loginMessage = resp.json()['loginMessage']
        if loginMessage == 'success':
            self.logger.info('Login successfully!')
            return True
        self.logger.error('Login failed as: {}'.format(loginMessage))
        return False

    def selectProjectInfosByCond(self):
        """
            项目列表
        """
        url = self.base_url + '/nhzy/project/project!selectProjectInfosByCond.action'

        formdata = {
            'cond.subprojectoperatorcode': self.username,
            'cond.kftype': 'A',
            'limit': 99999999,
            'start': 0,
            'page': 1,
        }
        resp = self.session.post(url, data=formdata).json()
        self.logger.info('Total Project Number: {}\n'.format(resp['projectInfos_num']))
        return resp['projectInfos']

    def selectSubprojectInfosByCond(self, projectnum):
        """
            子项目列表
        """
        url = self.base_url + '/nhzy/subproject/subproject!selectSubprojectInfosByCond.action'
        formdata = {
            'cond.projectnum': projectnum,
            'cond.batchsubprojectnum': 1,
            'limit': 99999999,
            'start': 0,
            'page': 1,
        }
        resp = self.session.post(url, data=formdata).json()
        return resp['subprojectInfos']

    def selectStageInfosByCond(self, subprojectnum):
        """
            分期查询
        """
        url = self.base_url + '/nhzy/settlementbill/stage!selectStageInfosByCond.action'
        formdata = {
            'cond.subprojectcode': subprojectnum,
            'limit': 99999999,
            'start': 0,
            'page': 1,
        }
        resp = self.session.post(url, data=formdata).json()

        return resp['stageInfos']

    def selectContractInfoByContractsno(self, contractno):
        """
            合同查询
        """
        url = self.base_url + '/crm/contract/contract!selectContractInfoByContractsno.action'
        formdata = {
            'cond.contractsno_equals': contractno,
        }
        resp = self.session.post(url, data=formdata).json()

        if len(resp['contractInfo']) != 1:
            self.logger.warn('not unique contract number: {}'.format(contractno))
            exit(1)

        return resp['contractInfo'][0]

    def selectQuotationproductInfosByCond(self, quotationid):
        """
            合同报价查询
        """
        url = self.base_url + '/nhzy/projectquotation/quotationproduct!selectQuotationproductInfosByCond.action'
        formdata = {
            'cond.kfquotationid': quotationid,
            'limit': 99999999,
            'start': 0,
            'page': 1,
        }
        resp = self.session.get(url, params=formdata).json()

        return resp['quotationproductInfos']

    def selectQuoprocessInfosByCond(self, kfquotationproductid):
        """
            工序查询
        """
        url = self.base_url + '/nhzy/projectquotation/quoprocess!selectQuoprocessInfosByCond.action'
        formdata = {
            'cond.kfquotationproductid': kfquotationproductid,
            'limit': 99999999,
            'start': 0,
            'page': 1,
        }
        resp = self.session.get(url, params=formdata).json()

        return resp['quoprocessInfos']

    def get_quotation_data(self, contractno):
        quotation_data = {}

        contract = self.selectContractInfoByContractsno(contractno)
        quotation_products = self.selectQuotationproductInfosByCond(contract['quotationid'])

        for quotation in quotation_products:
            pcode = quotation['pcode']

            if pcode not in quotation_data:
                quotation_data[pcode] = defaultdict(list)

            quotation_data[pcode]['samplenum'] = quotation['samplenum']
            quotation_data[pcode]['totalproductdata'] = 0

            # 工序列表
            quoprocesses = self.selectQuoprocessInfosByCond(quotation['kfquotationproductid'])
            for quoprocess in quoprocesses:
                if '上机' in quoprocess['processcode']:
                    quotation_data[pcode]['totalproductdata'] += float(quoprocess['datasize']) * int(quoprocess['samnum'])

                if quoprocess['datasize'] and quoprocess['datasize'].isdigit():
                    quoprocess['datasize'] = int(quoprocess['datasize'])
                else:
                    quoprocess['datasize'] = ''

                quotation_data[pcode]['processes'] += [
                    {
                        'processcode': quoprocess['processcode'],
                        'samnum': int(quoprocess['samnum']),
                        'datasize': quoprocess['datasize'],
                        'processtypename': quoprocess['processtypename']
                    }
                ]

            # quotation_data[pcode]['processnames'] += quotation['processnames'].split(',')
            quotation_data[pcode]['processtypename'] += quotation['processtypename'].split(',')
        return quotation_data


def main(**args):
    start_time = time.time()

    logger = get_logger('CMS', verbose=args['verbose'])

    logger.info('input arguments: {}'.format(args))

    # 结算样本信息
    report_data = parse_report_data(args['report'])

    # 配置文件解析
    config = parse_config(args['config'])
    if config:
        logger.debug('\033[32mConfig Data: \n{}\033[0m'.format(
            json.dumps(config, ensure_ascii=False, indent=2)))

    user = User(**args)
    username, password = user.get_user_pass()
    cms = CMS(username, password, logger=logger)
    if cms.login():
        user.save(username, password)
    else:
        exit(1)

    titles = [each[0] for each in title]
    for i in range(1, 6):
        stage_title[0] = '任务{} 任务名称'.format(i)
        titles += stage_title

    fields = [each[1] for each in title]

    book = openpyxl.Workbook()
    sheet1 = book.active
    sheet1.title = unicode('子项目')
    sheet2 = book.create_sheet(unicode('SOP明细'))

    # 写入表头
    write_title(sheet1, sheet2, titles)

    index1 = 0  # sheet1
    index2 = 1  # sheet2
    projects = cms.selectProjectInfosByCond()
    project_count = len(projects)

    for n, info in enumerate(projects, 1):
        # logger.info('>>> Projct: {projectnum}'.format(**info))

        percent = n * 100. / project_count
        sys.stderr.write(
            '\033[K[{n}/{project_count}] \033[36m{percent:.1f}% \033[0mcompleted\r'.format(**locals()))
        sys.stderr.flush()

        sub_projects = cms.selectSubprojectInfosByCond(info['projectnum'])

        # 项目 - 合同 - 工序
        quotation_data = cms.get_quotation_data(info['contractno'])

        for sub_info in sub_projects:  # 每个子项目编号为一行
            index1 += 1
            sub_info['index1'] = index1
            # logger.info('> Sub Project: {index1} {subprojectnum}'.format(**sub_info))

            sub_info['contractno'] = info['contractno']

            # 更新：config
            sub_info.update(config['common'])

            ave_sam_product, pooling = config['product'].get(unicode('{pname}'.format(**sub_info)))

            if sub_info['samplenum']:
                # ave_sam_product = sub_info.get(unicode('avesam{pname}'.format(**sub_info)))

                sub_info['pc_time'] = '=ROUNDUP(J{}/{},0) * {}'.format(index1+1,
                                                                       ave_sam_product, sub_info['pc_ave_time'])
                sub_info['bi_time'] = '=ROUNDUP(J{}/{},0) * {}'.format(index1+1,
                                                                       ave_sam_product, sub_info['bi_ave_time'])
                sub_info['pc_bi_time'] = '=U{0}+V{0}'.format(index1+1)

            # 更新：样本数，总数据量
            sub_info.update(quotation_data.get(sub_info['pcode'], {}))

            # 结算样本个数
            if report_data.get(sub_info['subprojectnum']):
                sub_info['jt_samplenum'] = report_data[sub_info['subprojectnum']]['samplenum']
            elif report_data.get((info['projectnum'], sub_info['pcode'])):
                result = report_data[(info['projectnum'], sub_info['pcode'])]
                sub_info['jt_samplenum'] = result['samplenum']
                if not sub_info.get('totalproductdata'):
                    sub_info['totalproductdata'] = result['datasize']

            # sub_info['jt_samplenum'] = report_data.get(sub_info['subprojectnum'], '')

            # SOP明细 -- 生产成本，不考虑分析
            # 测序前SOP成本 = SUM(样本数 * 对应单价)
            # 生产SOP成本 = 测序前SOP成本 + 数据量 * 单价 * pooling系数

            # pooling = sub_info.get(unicode('{pname}'.format(**sub_info)))

            sopcost_before = []
            sopcost_production = []

            if sub_info.get('processes'):
                for process in sub_info['processes']:
                    index2 += 1
                    sop_key = '{processcode}__{processtypename}'.format(**process)

                    sop_price = config['sop'].get(unicode(sop_key))

                    if sop_price:
                        logger.debug('{sop_key}: {sop_price}'.format(**locals()))

                        if '上机' in process['processcode']:
                            sopcost_production += ['K{}*{}*{}'.format(
                                index1+1, sop_price, pooling)]
                        elif any([each in process['processcode'] for each in ('提取', '库检', '检测', '建库')]):
                            sopcost_before += ['J{}*{}'.format(index1+1, sop_price)]

                    sheet2.cell(row=index2, column=1, value=sub_info['contractno'])
                    sheet2.cell(row=index2, column=2, value=sub_info['subprojectnum'])
                    sheet2.cell(row=index2, column=3, value=process['processcode'])
                    sheet2.cell(row=index2, column=4, value=process['processtypename'])
                    sheet2.cell(row=index2, column=5, value=process['samnum'])
                    sheet2.cell(row=index2, column=6, value=process['datasize'])

            if sopcost_before:
                sub_info['sopcost_before'] = '=' + ' + '.join(sopcost_before)
            if sopcost_production:
                sub_info['sopcost_production'] = '=L{} + '.format(
                    index1+1) + ' + '.join(sopcost_production)

            sub_info['compute_cost'] = '=K{i}*{gcluster}'.format(i=index1+1, **sub_info)
            sub_info['pc_bi_cost'] = '=U{i}*{pccost} + V{i}*{bicost}'.format(i=index1+1, **sub_info)
            sub_info['three_cost'] = '=I{i}/1.06*{apportion}'.format(i=index1+1, **sub_info)
            sub_info['profit'] = '=I{i}/1.06 - M{i} - N{i} - P{i} -Q{i}'.format(i=index1+1)

            # I/1.06 = 税后收入

            linelist = [sub_info.get(k, '') for k in fields]

            stages = cms.selectStageInfosByCond(sub_info['subprojectnum'])

            for stage_info in stages:
                tmp_list = [''] * 11
                tmp_list[0] = stage_info['installmentcode']
                tmp_list[1] = stage_info['operatemanagerdesc']
                tmp_list[3] = stage_info['informationleader']
                tmp_list[6] = stage_info['desc12']
                linelist += tmp_list

            for col, value in enumerate(linelist, 1):
                try:
                    value = float(value)
                except:
                    try:
                        value = int(value)
                    except:
                        pass

                if isinstance(value, list):
                    value = ','.join(value)
                sheet1.cell(row=index1 + 1, column=col, value=value)

        if args['limit'] and index1 >= args['limit']:
            break

    book.save(args['out'])

    sys.stderr.write('\n\n')
    logger.info('\033[33msave file: {out}\033[0m'.format(**args))
    logger.info('\033[32mtime used: {:.1f}s\033[0m'.format(time.time() - start_time))


def parser_add_cms(parser):

    parser.description = __doc__
    parser.epilog = textwrap.dedent('''
        \033[36mexamples:
            %(prog)s -o disease.xlsx
            %(prog)s -o disease.xlsx -r cms_report.xlsx      # 使用CMS导出的报表文件，用于更新结算样本数
            %(prog)s -o disease.xlsx -l 100                  # 只输出100条记录，可用于测试
            %(prog)s -o disease.xlsx -c your_config.xlsx     # 使用指定的配置文件
        \033[0m
    ''')

    parser.add_argument(
        '-c', '--config', help='the config excel [default: %(default)s]', default=DEFAULT_CONFIG)
    parser.add_argument(
        '-o', '--out', help='the output filename [default: %(default)s]', default='cms.project.xlsx')
    parser.add_argument(
        '-r', '--report', help='the report excel which contains sample informations')

    parser.add_argument('-l', '--limit', help='the limit count of output, just for test', type=int)

    parser.set_defaults(func=main)
