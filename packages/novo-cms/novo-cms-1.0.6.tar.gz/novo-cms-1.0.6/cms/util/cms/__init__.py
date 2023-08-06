#-*- encoding: utf8 -*-
import re
import warnings
from collections import defaultdict

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Color, colors, Alignment, Border, Side


warnings.simplefilter("ignore")


def write_title(sheet1, sheet2, titles):
    for col, value in enumerate(titles, 1):
        _ = sheet1.cell(row=1, column=col, value=value)
        if col in range(1, 9):
            color = '64F336'
        elif col in range(9, 24):
            color = 'EBBF1D'
        elif col in range(24, 36):
            color = '00FFF3'
        else:
            color = '3184EB'

        _.fill = PatternFill(start_color=color,
                             end_color=color,
                             fill_type="solid")
        _.font = Font(bold=True)
        if col > 1:
            sheet1.column_dimensions[get_column_letter(col)].width = 20

    color = 'FF8000'
    for col, value in enumerate(['合同编号', '子项目编号', '工序名称', '工序类型', '样本个数', '数据量'], 1):
        _ = sheet2.cell(row=1, column=col, value=value)
        _.font = Font(bold=True)
        _.fill = PatternFill(start_color=color,
                             end_color=color,
                             fill_type="solid")
        sheet2.column_dimensions[get_column_letter(col)].width = 24


def parse_report_data(report_xlsx):
    """
        从报表中获取结算样本信息
    """
    report_data = {}

    if report_xlsx:
        wb = openpyxl.load_workbook(report_xlsx)
        ws = wb.get_active_sheet()
        for row in ws.rows:
            if row[0].value == '序号':
                title = [each.value.decode() for each in row]
                if '子项目编号' in title:
                    key_indexes = [title.index('子项目编号')]
                    value_indexes = [title.index('结算样本个数')]
                else:
                    key_indexes = map(title.index, ['项目编号', '产品编码'])
                    value_indexes = map(title.index, ['样本数', '数据量'])
                continue
            elif not isinstance(row[0].value, float):
                continue

            keys = tuple(each.value for each in map(row.__getitem__, key_indexes))
            if len(keys) == 1:
                keys = keys[0]
            values = [each.value for each in map(row.__getitem__, value_indexes)]

            values = dict(zip(['samplenum', 'datasize'], values))

            if not(values['samplenum'] and values['samplenum'].isdigit()):
                continue

            if keys not in report_data:
                report_data[keys] = defaultdict(int)

            if values.get('samplenum'):
                report_data[keys]['samplenum'] += int(values['samplenum'])
            if values.get('datasize'):
                report_data[keys]['datasize'] += int(values['datasize'])

    # print report_data
    return report_data


def parse_config(configfile):
    wb = openpyxl.load_workbook(configfile)
    ws = wb.get_active_sheet()

    config = {
        'common': {},
        'product': {},
        'sop': {},
    }
    for row in ws.rows:

        if re.match(r'#|一级|工序名称', str(row[0].value)) or not row[0].value:
            continue

        if not row[1].value:
            config['common'][row[0].value] = row[2].value
        elif isinstance(row[1].value, long):
            config['product'][row[0].value] = [row[1].value, row[2].value]
        else:
            config['sop']['__'.join([row[0].value, row[1].value])] = row[2].value

    return config
